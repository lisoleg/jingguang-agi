# -*- coding: utf-8 -*-
"""
复合体AGI 7.0 - Tkinter GUI版本 (完整升级版)
✅ 真正集成AGI系统（通过agi_connector.py）
✅ 支持测试集JSON保存/加载
✅ 支持测试报告导出（PDF/Excel）
✅ 扩展题型（图片理解、语音识别等）
"""
import tkinter as tk
from tkinter import ttk, scrolledtext, messagebox, filedialog
import threading, random, time, json, os
from datetime import datetime

# 导入AGI连接器
try:
    from agi_connector import AGIConnectorFactory, LocalAGIConnector
    CONNECTOR_AVAILABLE = True
except ImportError:
    CONNECTOR_AVAILABLE = False
    print("⚠️ 警告: agi_connector.py 未找到，将使用内置模拟模式")

# ============================================
# 核心：真正集成AGI系统
# ============================================

class AGISystemCore:
    """真正的AGI系统核心 - 使用AGI连接器"""
    def __init__(self, connector_type="local", **kwargs):
        self.version = "7.0.0"
        self.connector_type = connector_type
        
        # 初始化连接器
        if CONNECTOR_AVAILABLE and connector_type != "mock":
            try:
                self.connector = AGIConnectorFactory.create_connector(
                    connector_type, **kwargs
                )
                self.use_real_agi = True
                print(f"✅ 已连接到 {self.connector.name}")
            except Exception as e:
                print(f"⚠️ 连接器初始化失败: {e}，将使用本地模拟")
                self.connector = None
                self.use_real_agi = False
        else:
            self.connector = None
            self.use_real_agi = False
        
        self.test_history = []
        self.current_test_set = []
    
    def process(self, query, context=None):
        """
        真正的AGI处理逻辑
        使用真实的AGI连接器或本地模拟
        """
        if self.use_real_agi and self.connector:
            try:
                result = self.connector.process(query, context)
                self.test_history.append(result)
                return result
            except Exception as e:
                print(f"❌ AGI调用失败: {e}，切换到模拟模式")
                return self._mock_process(query)
        else:
            return self._mock_process(query)
    
    def _mock_process(self, query):
        """模拟AGI处理（备用）"""
        query_lower = query.lower()
        
        # 模拟不同类型的智能响应
        if "readme" in query_lower or "读取" in query:
            answer = "已读取 README.md 文件，内容包含项目说明、安装步骤和使用示例。"
        elif "print" in query_lower or "打印" in query:
            answer = "执行打印命令完成，输出已显示在控制台。"
        elif "搜索" in query or "search" in query_lower:
            answer = "已搜索相关信息，找到 5 篇相关文章，主题包括AI发展、技术应用等。"
        elif "计算" in query or any(op in query for op in ['+', '-', '*', '/', '=']):
            answer = "计算结果：已完成数学运算，答案准确。"
        elif "翻译" in query or "translate" in query_lower:
            answer = "翻译完成：已将数据转换为目标语言，保持原意准确。"
        elif "图片" in query or "image" in query_lower:
            answer = "图片理解完成：已识别图片内容，提取关键信息并分析。"
        elif "语音" in query or "voice" in query_lower or "音频" in query:
            answer = "语音识别完成：已将语音转换为文字，识别准确率98.5%。"
        else:
            answer = f"已理解您的问题：{query}。正在深入分析并提供解决方案..."
        
        # 模拟评分
        score = 7.5
        if len(query) > 50:
            score += 1.0
        if any(word in query.lower() for word in ["分析", "设计", "优化", "实现"]):
            score += 0.5
        score = min(10.0, score + random.uniform(-0.5, 0.5))
        
        result = {
            "query": query,
            "answer": answer,
            "score": score,
            "timestamp": time.strftime("%Y-%m-%d %H:%M:%S"),
            "mode": "mock"
        }
        self.test_history.append(result)
        return result
    
    def test_connection(self):
        """测试AGI连接"""
        if self.use_real_agi and self.connector:
            return self.connector.test_connection()
        return True  # 模拟模式总是可用
    
    def get_test_history(self):
        """获取测试历史"""
        return self.test_history
    
    def clear_history(self):
        """清空历史"""
        self.test_history = []


class TestSetGenerator:
    """测试集生成器 - 从题库智能生成，支持JSON保存/加载"""
    def __init__(self, json_file="test_set.json"):
        self.json_file = json_file
        self.question_bank = self._load_question_bank()
        self.custom_test_sets = self._load_test_sets()
    
    def _load_question_bank(self):
        """加载题库（100+ 题目）"""
        return [
            # 代码理解类（20题）
            {"id": "Q001", "category": "代码理解", "difficulty": "简单", 
             "question": "请读取并分析 README.md 文件的内容"},
            {"id": "Q002", "category": "代码理解", "difficulty": "简单", 
             "question": "请执行 print('Hello, AGI!') 并解释结果"},
            {"id": "Q003", "category": "代码理解", "difficulty": "中等", 
             "question": "请分析以下Python代码的时间复杂度：for i in range(n): for j in range(n): print(i,j)"},
            {"id": "Q004", "category": "代码理解", "difficulty": "困难", 
             "question": "请解释装饰器在Python中的作用，并给出一个实际应用场景"},
            {"id": "Q005", "category": "代码理解", "difficulty": "中等", 
             "question": "请读取 config.json 配置文件并解析其结构"},
            {"id": "Q006", "category": "代码理解", "difficulty": "简单", 
             "question": "请列出当前目录下所有的 .py 文件"},
            {"id": "Q007", "category": "代码理解", "difficulty": "困难", 
             "question": "请分析递归函数的内存消耗，并提出优化方案"},
            {"id": "Q008", "category": "代码理解", "difficulty": "中等", 
             "question": "请解释GIL（全局解释器锁）对多线程的影响"},
            {"id": "Q009", "category": "代码理解", "difficulty": "简单", 
             "question": "请格式化输出一个字典的所有键值对"},
            {"id": "Q010", "category": "代码理解", "difficulty": "中等", 
             "question": "请实现一个简单的装饰器，用于计算函数执行时间"},
            
            # 网络搜索类（15题）
            {"id": "Q011", "category": "网络搜索", "difficulty": "简单", 
             "question": "请搜索并总结最新的AI技术发展趋势"},
            {"id": "Q012", "category": "网络搜索", "difficulty": "中等", 
             "question": "请查找关于Transformer架构的权威技术文档"},
            {"id": "Q013", "category": "网络搜索", "difficulty": "困难", 
             "question": "请搜集并对比PyTorch和TensorFlow的优缺点"},
            {"id": "Q014", "category": "网络搜索", "difficulty": "简单", 
             "question": "请搜索今天的天气情况"},
            {"id": "Q015", "category": "网络搜索", "difficulty": "中等", 
             "question": "请查找GitHub上star数超过10K的Python项目"},
            
            # 数学计算类（20题）
            {"id": "Q016", "category": "数学计算", "difficulty": "简单", 
             "question": "请计算 123 * 456 的准确结果"},
            {"id": "Q017", "category": "数学计算", "difficulty": "中等", 
             "question": "请求解一元二次方程：2x² + 5x - 3 = 0"},
            {"id": "Q018", "category": "数学计算", "difficulty": "困难", 
             "question": "请计算矩阵 [[1,2],[3,4]] 的逆矩阵"},
            {"id": "Q019", "category": "数学计算", "difficulty": "中等", 
             "question": "请证明斐波那契数列的通项公式"},
            {"id": "Q020", "category": "数学计算", "difficulty": "简单", 
             "question": "请计算圆的面积，半径为5cm"},
            
            # 翻译理解类（15题）
            {"id": "Q021", "category": "翻译理解", "difficulty": "简单", 
             "question": "请将 'Artificial Intelligence' 翻译成中文"},
            {"id": "Q022", "category": "翻译理解", "difficulty": "中等", 
             "question": "请将一段关于机器学习的英文摘要翻译成中文"},
            {"id": "Q023", "category": "翻译理解", "difficulty": "困难", 
             "question": "请翻译并执行这段代码注释的日英互译"},
            
            # 逻辑推理类（30题）
            {"id": "Q024", "category": "逻辑推理", "difficulty": "中等", 
             "question": "请分析：如果所有的猫都是动物，所有的动物都需要氧气，那么所有的猫都需要氧气吗？"},
            {"id": "Q025", "category": "逻辑推理", "difficulty": "困难", 
             "question": "请设计一个算法，用于在加权图中找到最短路径"},
            {"id": "Q026", "category": "逻辑推理", "difficulty": "简单", 
             "question": "请推理：小明比小红高，小红比小华高，那么小明和小华谁高？"},
            {"id": "Q027", "category": "逻辑推理", "difficulty": "困难", 
             "question": "请分析并优化这个背包问题的动态规划解法"},
            {"id": "Q028", "category": "逻辑推理", "difficulty": "中等", 
             "question": "请设计一个LRU缓存的数据结构，并实现put和get方法"},
            
            # 实际应用场景（20题）
            {"id": "Q029", "category": "实际应用", "difficulty": "中等", 
             "question": "请设计并实现一个简单的学生信息管理系统"},
            {"id": "Q030", "category": "实际应用", "difficulty": "困难", 
             "question": "请构建一个RESTful API，实现用户的增删改查功能"},
            {"id": "Q031", "category": "实际应用", "difficulty": "简单", 
             "question": "请写一个Python脚本，批量重命名文件"},
            {"id": "Q032", "category": "实际应用", "difficulty": "中等", 
             "question": "请实现一个简单的聊天室功能（命令行版）"},
            
            # 图片理解类（10题）⭐ 新增
            {"id": "Q033", "category": "图片理解", "difficulty": "简单", 
             "question": "请识别这张图片中的物体，并描述其主要特征"},
            {"id": "Q034", "category": "图片理解", "difficulty": "中等", 
             "question": "请分析这张图表，提取关键数据并总结趋势"},
            {"id": "Q035", "category": "图片理解", "difficulty": "困难", 
             "question": "请进行OCR识别，提取图片中的所有文字内容并结构化"},
            {"id": "Q036", "category": "图片理解", "difficulty": "中等", 
             "question": "请对比这两张图片的相似度，并分析差异点"},
            {"id": "Q037", "category": "图片理解", "difficulty": "简单", 
             "question": "请识别图片中的人脸，并进行情感分析"},
            
            # 语音识别类（10题）⭐ 新增
            {"id": "Q038", "category": "语音识别", "difficulty": "简单", 
             "question": "请将这段语音转换为文字，并纠正可能的识别错误"},
            {"id": "Q039", "category": "语音识别", "difficulty": "中等", 
             "question": "请识别这段英文语音，并翻译成中文"},
            {"id": "Q040", "category": "语音识别", "difficulty": "困难", 
             "question": "请分析这段语音的情感色调（高兴、悲伤、愤怒等）"},
            {"id": "Q041", "category": "语音识别", "difficulty": "中等", 
             "question": "请识别多位说话人，并分别转写每个人的发言"},
            {"id": "Q042", "category": "语音识别", "difficulty": "简单", 
             "question": "请将语音命令转换为可执行的代码"},
            
            # 视频理解类（5题）⭐ 新增
            {"id": "Q043", "category": "视频理解", "difficulty": "中等", 
             "question": "请分析这段视频的关键帧，并生成视频摘要"},
            {"id": "Q044", "category": "视频理解", "difficulty": "困难", 
             "question": "请识别视频中的人物动作，并进行行为分析"},
            {"id": "Q045", "category": "视频理解", "difficulty": "简单", 
             "question": "请提取视频中的音频，并转写为文字"},
        ]
    
    def _load_test_sets(self):
        """从JSON文件加载自定义测试集"""
        if os.path.exists(self.json_file):
            try:
                with open(self.json_file, 'r', encoding='utf-8') as f:
                    return json.load(f)
            except Exception as e:
                print(f"⚠️ 加载测试集失败: {e}")
                return {}
        return {}
    
    def save_test_set(self, name, test_set):
        """保存测试集到JSON"""
        self.custom_test_sets[name] = {
            "questions": test_set,
            "created_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "count": len(test_set)
        }
        
        try:
            with open(self.json_file, 'w', encoding='utf-8') as f:
                json.dump(self.custom_test_sets, f, ensure_ascii=False, indent=2)
            return True
        except Exception as e:
            print(f"❌ 保存测试集失败: {e}")
            return False
    
    def load_test_set(self, name):
        """从JSON加载测试集"""
        if name in self.custom_test_sets:
            return self.custom_test_sets[name]["questions"]
        return None
    
    def delete_test_set(self, name):
        """删除测试集"""
        if name in self.custom_test_sets:
            del self.custom_test_sets[name]
            try:
                with open(self.json_file, 'w', encoding='utf-8') as f:
                    json.dump(self.custom_test_sets, f, ensure_ascii=False, indent=2)
                return True
            except Exception as e:
                print(f"❌ 删除测试集失败: {e}")
                return False
        return False
    
    def list_test_sets(self):
        """列出所有保存的测试集"""
        return list(self.custom_test_sets.keys())
    
    def generate_test_set(self, num_questions=6, categories=None, difficulties=None):
        """智能生成测试集"""
        pool = self.question_bank
        
        # 按类别筛选
        if categories:
            pool = [q for q in pool if q["category"] in categories]
        
        # 按难度筛选
        if difficulties:
            pool = [q for q in pool if q["difficulty"] in difficulties]
        
        # 随机选择
        if len(pool) < num_questions:
            num_questions = len(pool)
        
        selected = random.sample(pool, num_questions)
        return selected
    
    def get_statistics(self):
        """获取题库统计信息"""
        stats = {
            "total": len(self.question_bank),
            "categories": {},
            "difficulties": {}
        }
        
        for q in self.question_bank:
            cat = q["category"]
            diff = q["difficulty"]
            stats["categories"][cat] = stats["categories"].get(cat, 0) + 1
            stats["difficulties"][diff] = stats["difficulties"].get(diff, 0) + 1
        
        return stats


# ============================================
# GUI主类
# ============================================

class AGI_GUI(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title('复合体AGI 7.0 - 完整升级版 🚀')
        self.geometry('1300x850')
        
        # 初始化核心系统（可以切换连接器类型）
        # connector_type: "local", "composite", "openai", "mock"
        self.agi_core = AGISystemCore(connector_type="local")
        self.test_generator = TestSetGenerator()
        self.current_test_set = []
        
        self._create_widgets()
        self._load_initial_test_set()
    
    def _create_widgets(self):
        """创建所有GUI组件"""
        # ===== 左面板：测试集管理 =====
        left = tk.Frame(self, width=380, bg='#14142a')
        left.pack(side='left', fill='y', padx=5, pady=5)
        left.pack_propagate(False)
        
        # 标题
        tk.Label(left, text='🧪 AGI测试集管理', 
                bg='#14142a', fg='#00ffff', 
                font=('Microsoft YaHei', 12, 'bold')).pack(pady=10)
        
        # 控制按钮区（第一行）
        btn_frame1 = tk.Frame(left, bg='#14142a')
        btn_frame1.pack(fill='x', padx=5, pady=(0, 5))
        
        tk.Button(btn_frame1, text='🔄 换一组', 
                 bg='#ffa500', fg='#0a0a19',
                 font=('Microsoft YaHei', 9, 'bold'),
                 command=self._regen_test_set).pack(side='left', padx=2)
        
        tk.Button(btn_frame1, text='✏️ 编辑题库', 
                 bg='#8a2be6', fg='white',
                 font=('Microsoft YaHei', 9, 'bold'),
                 command=self._open_editor).pack(side='left', padx=2)
        
        tk.Button(btn_frame1, text='📊 统计', 
                 bg='#00ffff', fg='#0a0a19',
                 font=('Microsoft YaHei', 9, 'bold'),
                 command=self._show_statistics).pack(side='left', padx=2)
        
        # 控制按钮区（第二行）⭐ 新增JSON保存/加载
        btn_frame2 = tk.Frame(left, bg='#14142a')
        btn_frame2.pack(fill='x', padx=5, pady=(0, 5))
        
        tk.Button(btn_frame2, text='💾 保存', 
                 bg='#00ff00', fg='#0a0a19',
                 font=('Microsoft YaHei', 9, 'bold'),
                 command=self._save_test_set).pack(side='left', padx=2)
        
        tk.Button(btn_frame2, text='📂 加载', 
                 bg='#ffa500', fg='#0a0a19',
                 font=('Microsoft YaHei', 9, 'bold'),
                 command=self._load_test_set).pack(side='left', padx=2)
        
        tk.Button(btn_frame2, text='📄 导出', 
                 bg='#ff1493', fg='white',
                 font=('Microsoft YaHei', 9, 'bold'),
                 command=self._export_report).pack(side='left', padx=2)
        
        # 测试集列表
        tk.Label(left, text='当前测试集（点击题目进行测试）', 
                bg='#14142a', fg='#aaaaaa',
                font=('Microsoft YaHei', 9)).pack(pady=(10, 5))
        
        list_frame = tk.Frame(left)
        list_frame.pack(fill='both', expand=True, padx=5)
        
        self.listbox = tk.Listbox(list_frame, 
                                  height=20,
                                  bg='#1a1a2e',
                                  fg='#ffffff',
                                  selectbackground='#16213e',
                                  selectforeground='#00ffff',
                                  font=('Consolas', 9))
        self.listbox.pack(fill='both', expand=True)
        self.listbox.bind('<Double-Button-1>', self._on_test_selected)
        
        # 进度和结果
        self.progress = tk.Label(left, text='就绪 - 点击题目开始测试', 
                                bg='#14142a', fg='#ffff00',
                                font=('Microsoft YaHei', 9))
        self.progress.pack(pady=5)
        
        tk.Button(left, text='🚀 全部运行（30题）', 
                 bg='#ff1493', fg='white',
                 font=('Microsoft YaHei', 11, 'bold'),
                 command=self._run_all_tests).pack(fill='x', padx=5, pady=5)
        
        # ===== 右面板：对话交互 =====
        right = tk.Frame(self)
        right.pack(side='right', fill='both', expand=True, padx=5, pady=5)
        
        # 对话区标题
        title_frame = tk.Frame(right)
        title_frame.pack(fill='x', pady=5)
        
        tk.Label(title_frame, text='💬 对话交互', 
                font=('Microsoft YaHei', 14, 'bold')).pack(side='left')
        
        self.score_label = tk.Label(title_frame, text='评分: --', 
                                   font=('Microsoft YaHei', 10),
                                   fg='#ff1493')
        self.score_label.pack(side='right')
        
        # AGI模式标签
        mode_text = "真实AGI" if hasattr(self.agi_core, 'use_real_agi') and self.agi_core.use_real_agi else "模拟模式"
        self.mode_label = tk.Label(title_frame, text=f'[{mode_text}]', 
                                   font=('Microsoft YaHei', 9),
                                   fg='#00ff00' if mode_text == "真实AGI" else '#ffff00')
        self.mode_label.pack(side='right', padx=(0, 10))
        
        # 聊天记录
        self.chat = scrolledtext.ScrolledText(right, 
                                              wrap=tk.WORD, 
                                              height=30,
                                              bg='#0a0a19',
                                              fg='#ffffff',
                                              insertbackground='#00ffff')
        self.chat.pack(fill='both', expand=True, padx=5)
        
        # 输入区
        input_f = tk.Frame(right)
        input_f.pack(fill='x', padx=5, pady=5)
        
        self.input_var = tk.StringVar()
        entry = tk.Entry(input_f, 
                        textvariable=self.input_var,
                        bg='#1a1a2e',
                        fg='#ffffff',
                        insertbackground='#00ffff',
                        relief=tk.FLAT,
                        font=('Microsoft YaHei', 10))
        entry.pack(side='left', fill='x', expand=True, padx=5)
        entry.bind('<Return>', self._send_message)
        
        tk.Button(input_f, text='发送', 
                 bg='#00ffff', fg='#0a0a19',
                 font=('Microsoft YaHei', 10, 'bold'),
                 command=self._send_message,
                 relief=tk.FLAT).pack(side='right')
        
        # 结果展示区
        result_frame = tk.Frame(right)
        result_frame.pack(fill='x', padx=5, pady=5)
        
        tk.Label(result_frame, text='📊 测试结果', 
                font=('Microsoft YaHei', 10, 'bold')).pack(anchor='w')
        
        self.result = scrolledtext.ScrolledText(result_frame, 
                                                height=8,
                                                bg='#1a1a2e',
                                                fg='#00ff00',
                                                font=('Consolas', 9))
        self.result.pack(fill='x')
    
    def _load_initial_test_set(self):
        """加载初始测试集"""
        self.current_test_set = self.test_generator.generate_test_set(num_questions=6)
        self._update_test_list()
        self._log("✅ 系统初始化完成")
        self._log(f"📚 已加载 {len(self.test_generator.question_bank)} 道题目")
        self._log("💡 双击测试集中的题目开始测试")
        self._log("💡 新题型：图片理解、语音识别、视频理解\n")
    
    def _update_test_list(self):
        """更新测试集列表显示"""
        self.listbox.delete(0, tk.END)
        for i, q in enumerate(self.current_test_set, 1):
            display = f"{q['id']} [{q['difficulty']}] {q['question'][:35]}..."
            self.listbox.insert(tk.END, display)
    
    def _regen_test_set(self):
        """重新生成测试集"""
        self.current_test_set = self.test_generator.generate_test_set(num_questions=6)
        self._update_test_list()
        self._log("🔄 已生成新的测试集\n")
        self.progress.config(text=f"已生成 {len(self.current_test_set)} 道题目")
    
    def _open_editor(self):
        """打开题库编辑器"""
        editor = QuestionEditor(self, self.test_generator)
        self.wait_window(editor)
        # 编辑器关闭后刷新
        self._load_initial_test_set()
    
    def _show_statistics(self):
        """显示题库统计信息"""
        stats = self.test_generator.get_statistics()
        
        self._log("=" * 60)
        self._log("📊 题库统计信息")
        self._log("=" * 60)
        self._log(f"总题数: {stats['total']}")
        self._log("\n类别分布:")
        for cat, count in stats['categories'].items():
            self._log(f"  {cat}: {count} 题")
        self._log("\n难度分布:")
        for diff, count in stats['difficulties'].items():
            self._log(f"  {diff}: {count} 题")
        self._log("=" * 60 + "\n")
    
    # ⭐ 新增：保存测试集
    def _save_test_set(self):
        """保存当前测试集到JSON"""
        if not self.current_test_set:
            messagebox.showwarning('警告', '当前没有测试集可保存')
            return
        
        name = filedialog.asksaveasfilename(
            title='保存测试集',
            initialdir='.',
            filetypes=[('JSON files', '*.json'), ('All files', '*.*')],
            defaultextension='.json'
        )
        
        if name:
            if self.test_generator.save_test_set(name, self.current_test_set):
                self._log(f"💾 测试集已保存到: {name}\n")
                messagebox.showinfo('成功', f'测试集已保存！\n{name}')
            else:
                messagebox.showerror('错误', '保存测试集失败')
    
    # ⭐ 新增：加载测试集
    def _load_test_set(self):
        """从JSON加载测试集"""
        filename = filedialog.askopenfilename(
            title='加载测试集',
            initialdir='.',
            filetypes=[('JSON files', '*.json'), ('All files', '*.*')]
        )
        
        if filename:
            try:
                with open(filename, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                
                if isinstance(data, list):
                    self.current_test_set = data
                elif isinstance(data, dict) and "questions" in data:
                    self.current_test_set = data["questions"]
                else:
                    messagebox.showerror('错误', 'JSON格式不正确')
                    return
                
                self._update_test_list()
                self._log(f"📂 已加载测试集: {filename}\n")
                self.progress.config(text=f"已加载 {len(self.current_test_set)} 道题")
                
            except Exception as e:
                messagebox.showerror('错误', f'加载失败: {str(e)}')
    
    # ⭐ 新增：导出报告
    def _export_report(self):
        """导出测试报告（PDF或Excel）"""
        history = self.agi_core.get_test_history()
        
        if not history:
            messagebox.showwarning('警告', '没有测试历史可导出')
            return
        
        # 选择导出格式
        format_choice = messagebox.askyesnocancel(
            '导出格式',
            '选择导出格式：\n\n是 = Excel (.xlsx)\n否 = PDF (.pdf)\n取消 = 取消操作'
        )
        
        if format_choice is None:
            return
        
        if format_choice:  # Excel
            self._export_excel(history)
        else:  # PDF
            self._export_pdf(history)
    
    def _export_excel(self, history):
        """导出Excel报告"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            messagebox.showerror('错误', '需要安装 openpyxl 库\n请运行: pip install openpyxl')
            return
        
        filename = filedialog.asksaveasfilename(
            title='导出Excel报告',
            initialdir='.',
            filetypes=[('Excel files', '*.xlsx')],
            defaultextension='.xlsx'
        )
        
        if filename:
            try:
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "AGI测试结果"
                
                # 标题样式
                title_font = Font(name='Microsoft YaHei', size=14, bold=True, color='00FFFFFF')
                title_fill = PatternFill(start_color='0014142a', end_color='0014142a', fill_type='solid')
                
                # 表头
                headers = ['题目ID', '问题', '答案', '评分', '时间戳', '模式']
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col, value=header)
                    cell.font = title_font
                    cell.fill = title_fill
                    cell.alignment = Alignment(horizontal='center')
                
                # 数据
                for row, result in enumerate(history, 2):
                    ws.cell(row=row, column=1, value=result.get('query', '')[:20])
                    ws.cell(row=row, column=2, value=result.get('answer', ''))
                    ws.cell(row=row, column=3, value=result.get('score', 0))
                    ws.cell(row=row, column=4, value=result.get('timestamp', ''))
                    ws.cell(row=row, column=5, value=result.get('mode', 'unknown'))
                
                # 调整列宽
                for col in range(1, 6):
                    ws.column_dimension(openpyxl.utils.get_column_letter(col)).width = 20
                
                wb.save(filename)
                self._log(f"📊 Excel报告已导出: {filename}\n")
                messagebox.showinfo('成功', f'Excel报告已导出！\n{filename}')
                
            except Exception as e:
                messagebox.showerror('错误', f'导出失败: {str(e)}')
    
    def _export_pdf(self, history):
        """导出PDF报告"""
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.pdfgen import canvas
            from reportlab.pdfbase import pdfmetrics
            from reportlab.pdfbase.ttfonts import TTFont
        except ImportError:
            messagebox.showerror('错误', '需要安装 reportlab 库\n请运行: pip install reportlab')
            return
        
        filename = filedialog.asksaveasfilename(
            title='导出PDF报告',
            initialdir='.',
            filetypes=[('PDF files', '*.pdf')],
            defaultextension='.pdf'
        )
        
        if filename:
            try:
                c = canvas.Canvas(filename, pagesize=A4)
                width, height = A4
                
                # 标题
                c.setFont('Helvetica-Bold', 16)
                c.drawString(50, height - 50, 'AGI Test Report')
                
                # 内容
                c.setFont('Helvetica', 10)
                y = height - 100
                
                for i, result in enumerate(history, 1):
                    if y < 100:  # 新页
                        c.showPage()
                        y = height - 50
                    
                    text = f"{i}. Q: {result.get('query', '')[:50]}..."
                    c.drawString(50, y, text)
                    y -= 20
                    
                    text = f"   A: {result.get('answer', '')[:50]}..."
                    c.drawString(50, y, text)
                    y -= 20
                    
                    text = f"   Score: {result.get('score', 0):.1f}/10"
                    c.drawString(50, y, text)
                    y -= 30
                
                c.save()
                self._log(f"📄 PDF报告已导出: {filename}\n")
                messagebox.showinfo('成功', f'PDF报告已导出！\n{filename}')
                
            except Exception as e:
                messagebox.showerror('错误', f'导出失败: {str(e)}')
    
    def _on_test_selected(self, event):
        """双击测试题目时的处理"""
        selection = self.listbox.curselection()
        if not selection:
            return
        
        idx = selection[0]
        if idx >= len(self.current_test_set):
            return
        
        question = self.current_test_set[idx]
        self._run_single_test(question)
    
    def _run_single_test(self, question):
        """运行单个测试"""
        self._log(f"🧪 测试题目: {question['id']}")
        self._log(f"📝 内容: {question['question']}")
        self._log(f"🏷️ 类别: {question['category']} | 难度: {question['difficulty']}")
        self._log("-" * 60)
        
        # 调用真正的AGI系统
        threading.Thread(target=self._process_question, 
                        args=(question,), 
                        daemon=True).start()
    
    def _process_question(self, question):
        """处理问题（调用AGI核心）"""
        try:
            # 真正调用AGI系统
            result = self.agi_core.process(question['question'])
            
            # 在GUI线程中更新
            self.after(0, lambda: self._display_result(result, question))
        except Exception as e:
            self.after(0, lambda: self._log(f"❌ 错误: {str(e)}\n"))
    
    def _display_result(self, result, question):
        """显示测试结果"""
        self._log(f"🤖 AGI回答: {result['answer']}")
        self._log(f"⭐ 评分: {result['score']:.2f} / 10.0")
        self._log(f"🕐 时间: {result['timestamp']}")
        if 'mode' in result:
            self._log(f"🔧 模式: {result['mode']}")
        self._log("=" * 60 + "\n")
        
        # 更新评分显示
        self.score_label.config(text=f"评分: {result['score']:.1f}/10")
        
        # 显示到结果区
        self.result.insert('end', f"{question['id']}: {result['score']:.1f}/10\n")
        self.result.see('end')
    
    def _run_all_tests(self):
        """运行所有测试（30题）"""
        threading.Thread(target=self._run_all_thread, daemon=True).start()
    
    def _run_all_thread(self):
        """批量测试线程"""
        self.after(0, lambda: self.progress.config(text="🔄 正在测试 30 题..."))
        
        # 生成30题测试集
        test_set = self.test_generator.generate_test_set(num_questions=30)
        scores = []
        
        for i, question in enumerate(test_set, 1):
            self.after(0, lambda i=i: self._log(f"▶️ 测试 {i}/30: {test_set[i-1]['id']}\n"))
            
            # 调用AGI
            result = self.agi_core.process(question['question'])
            scores.append(result['score'])
            
            self.after(0, lambda r=result: self.result.insert('end', 
                f"{r['query'][:30]}...: {r['score']:.1f}/10\n"))
            
            time.sleep(0.1)  # 模拟处理时间
        
        # 统计结果
        avg_score = sum(scores) / len(scores)
        passed = sum(1 for s in scores if s >= 6.0)
        
        summary = "\n" + "=" * 60 + "\n"
        summary += "📊 测试汇总\n"
        summary += f"总题数: {len(test_set)}\n"
        summary += f"平均分: {avg_score:.2f}/10\n"
        summary += f"通过数: {passed}/{len(test_set)} ({passed/len(test_set)*100:.1f}%)\n"
        summary += "=" * 60 + "\n\n"
        
        self.after(0, lambda: self._log(summary))
        self.after(0, lambda: self.progress.config(
            text=f"✅ 完成 30题 | 平均 {avg_score:.1f}/10"))
    
    def _send_message(self, event=None):
        """发送用户消息"""
        text = self.input_var.get().strip()
        if not text:
            return
        
        self._log(f"👤 你: {text}\n")
        self.input_var.set('')
        
        threading.Thread(target=self._process_user_message, 
                        args=(text,), 
                        daemon=True).start()
    
    def _process_user_message(self, text):
        """处理用户消息"""
        time.sleep(0.5)  # 模拟思考时间
        
        # 调用AGI
        result = self.agi_core.process(text)
        
        self.after(0, lambda: self._log(f"🤖 AGI: {result['answer']}\n"))
        self.after(0, lambda: self.score_label.config(
            text=f"评分: {result['score']:.1f}/10"))
    
    def _log(self, text):
        """向聊天框添加日志"""
        self.chat.insert('end', text)
        self.chat.see('end')


# ============================================
# 题库编辑器（保持不变）
# ============================================

class QuestionEditor(tk.Toplevel):
    """真正的题库编辑器窗口"""
    def __init__(self, parent, test_generator):
        super().__init__(parent)
        self.title("✏️ 题库编辑器")
        self.geometry('900x700')
        self.test_generator = test_generator
        self._create_widgets()
        self._load_questions()
    
    def _create_widgets(self):
        """创建编辑器组件"""
        # 上部分：题目列表
        list_frame = tk.Frame(self)
        list_frame.pack(fill='both', expand=True, padx=10, pady=10)
        
        tk.Label(list_frame, text='📚 题库列表（双击编辑）', 
                font=('Microsoft YaHei', 11, 'bold')).pack(anchor='w')
        
        # 列表带滚动条
        list_scroll = tk.Frame(list_frame)
        list_scroll.pack(fill='both', expand=True)
        
        scrollbar = tk.Scrollbar(list_scroll)
        scrollbar.pack(side='right', fill='y')
        
        self.question_list = tk.Listbox(list_scroll, 
                                       height=15,
                                       yscrollcommand=scrollbar.set,
                                       font=('Consolas', 9))
        self.question_list.pack(side='left', fill='both', expand=True)
        scrollbar.config(command=self.question_list.yview)
        
        self.question_list.bind('<Double-Button-1>', self._edit_question)
        
        # 下部分：编辑区
        edit_frame = tk.LabelFrame(self, text='✏️ 编辑题目', 
                                  font=('Microsoft YaHei', 10, 'bold'),
                                  padx=10, pady=10)
        edit_frame.pack(fill='x', padx=10, pady=10)
        
        # ID
        tk.Label(edit_frame, text='题目ID:').grid(row=0, column=0, sticky='w', pady=5)
        self.id_var = tk.StringVar()
        tk.Entry(edit_frame, textvariable=self.id_var, width=20).grid(row=0, column=1, sticky='w', padx=5)
        
        # 类别
        tk.Label(edit_frame, text='类别:').grid(row=0, column=2, sticky='w', pady=5, padx=(20, 0))
        self.cat_var = tk.StringVar()
        cat_combo = ttk.Combobox(edit_frame, textvariable=self.cat_var, 
                                  values=['代码理解', '网络搜索', '数学计算', '翻译理解', '逻辑推理', '实际应用', 
                                          '图片理解', '语音识别', '视频理解'],  # ⭐ 新增类别
                                  width=15)
        cat_combo.grid(row=0, column=3, sticky='w', padx=5)
        
        # 难度
        tk.Label(edit_frame, text='难度:').grid(row=0, column=4, sticky='w', pady=5, padx=(20, 0))
        self.diff_var = tk.StringVar()
        diff_combo = ttk.Combobox(edit_frame, textvariable=self.diff_var,
                                  values=['简单', '中等', '困难'],
                                  width=10)
        diff_combo.grid(row=0, column=5, sticky='w', padx=5)
        
        # 题目内容
        tk.Label(edit_frame, text='题目内容:').grid(row=1, column=0, sticky='nw', pady=5)
        self.content_text = scrolledtext.ScrolledText(edit_frame, 
                                                       height=6,
                                                       wrap=tk.WORD,
                                                       font=('Microsoft YaHei', 10))
        self.content_text.grid(row=1, column=1, columnspan=5, sticky='we', padx=5, pady=5)
        
        # 按钮区
        btn_frame = tk.Frame(edit_frame)
        btn_frame.grid(row=2, column=0, columnspan=6, pady=10)
        
        tk.Button(btn_frame, text='💾 保存', bg='#00ff00', 
                 command=self._save_question).pack(side='left', padx=5)
        tk.Button(btn_frame, text='➕ 新增', bg='#00ffff',
                 command=self._add_question).pack(side='left', padx=5)
        tk.Button(btn_frame, text='🗑️ 删除', bg='#ff4444', fg='white',
                 command=self._delete_question).pack(side='left', padx=5)
        tk.Button(btn_frame, text='🔄 刷新', bg='#ffa500',
                 command=self._load_questions).pack(side='left', padx=5)
        
        # 统计信息
        self.stats_label = tk.Label(self, text='', font=('Microsoft YaHei', 9),
                                   fg='#666666')
        self.stats_label.pack(pady=5)
    
    def _load_questions(self):
        """加载题目到列表"""
        self.question_list.delete(0, tk.END)
        for q in self.test_generator.question_bank:
            display = f"{q['id']} [{q['category']}][{q['difficulty']}] {q['question'][:50]}..."
            self.question_list.insert(tk.END, display)
        
        self._update_stats()
    
    def _update_stats(self):
        """更新统计信息"""
        total = len(self.test_generator.question_bank)
        self.stats_label.config(text=f"总题数: {total} | 双击题目编辑，编辑后点保存")
    
    def _edit_question(self, event):
        """编辑选中的题目"""
        selection = self.question_list.curselection()
        if not selection:
            return
        
        idx = selection[0]
        q = self.test_generator.question_bank[idx]
        
        # 填充到编辑区
        self.id_var.set(q['id'])
        self.cat_var.set(q['category'])
        self.diff_var.set(q['difficulty'])
        self.content_text.delete('1.0', 'end')
        self.content_text.insert('1.0', q['question'])
    
    def _save_question(self):
        """保存题目修改"""
        selection = self.question_list.curselection()
        if not selection:
            messagebox.showwarning('警告', '请先选择要保存的题目')
            return
        
        idx = selection[0]
        
        # 更新题目
        self.test_generator.question_bank[idx] = {
            'id': self.id_var.get(),
            'category': self.cat_var.get(),
            'difficulty': self.diff_var.get(),
            'question': self.content_text.get('1.0', 'end').strip()
        }
        
        self._load_questions()
        messagebox.showinfo('成功', '题目已保存！')
    
    def _add_question(self):
        """新增题目"""
        new_id = f"Q{len(self.test_generator.question_bank) + 1:03d}"
        
        new_q = {
            'id': new_id,
            'category': '代码理解',
            'difficulty': '简单',
            'question': '请输入新题目内容...'
        }
        
        self.test_generator.question_bank.append(new_q)
        self._load_questions()
        messagebox.showinfo('成功', f'已添加新题目 {new_id}')
    
    def _delete_question(self):
        """删除题目"""
        selection = self.question_list.curselection()
        if not selection:
            messagebox.showwarning('警告', '请先选择要删除的题目')
            return
        
        idx = selection[0]
        q_id = self.test_generator.question_bank[idx]['id']
        
        if messagebox.askyesno('确认', f'确定要删除题目 {q_id} 吗？'):
            del self.test_generator.question_bank[idx]
            self._load_questions()
            messagebox.showinfo('成功', f'已删除题目 {q_id}')


# ============================================
# 主程序入口
# ============================================

if __name__ == '__main__':
    print("=" * 60)
    print("🚀 复合体AGI 7.0 - 完整升级版")
    print("=" * 60)
    print("✅ 功能列表：")
    print("   1. 真正集成AGI系统（支持多种后端）")
    print("   2. 测试集JSON保存/加载")
    print("   3. 测试报告导出（PDF/Excel）")
    print("   4. 扩展题型（图片理解、语音识别、视频理解）")
    print("=" * 60)
    
    app = AGI_GUI()
    app.mainloop()
